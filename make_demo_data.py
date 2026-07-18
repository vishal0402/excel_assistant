"""Generates demo_academic_data.xlsx — sample data for the AI Excel Assistant demo."""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(7)

FIRST = ["Aarav", "Vivaan", "Aditya", "Ishaan", "Kabir", "Ananya", "Diya", "Isha",
          "Myra", "Saanvi", "Reyansh", "Arjun", "Kiara", "Anaya", "Vihaan", "Rohan",
          "Meera", "Priya", "Karan", "Neha", "Aryan", "Riya", "Yash", "Tara",
          "Devansh", "Pooja", "Nikhil", "Simran", "Rahul", "Sneha"]
LAST = ["Sharma", "Verma", "Patel", "Gupta", "Reddy", "Nair", "Iyer", "Singh",
         "Mehta", "Chawla", "Kapoor", "Joshi", "Rao", "Bhat", "Kulkarni"]

SUBJECTS = ["Data Structures", "DBMS", "Operating Systems", "Computer Networks",
            "Software Engineering"]
DEPARTMENTS = ["CSE", "IT", "ECE"]

rows = []
roll_start = 2201

names_used = set()

def unique_name():
    while True:
        name = f"{random.choice(FIRST)} {random.choice(LAST)}"
        if name not in names_used:
            names_used.add(name)
            return name

n_students = 55
for i in range(n_students):
    name = unique_name()
    roll = f"CS{roll_start + i}"
    dept = random.choice(DEPARTMENTS)
    subject = random.choice(SUBJECTS)
    marks = random.randint(28, 100)
    attendance = random.randint(48, 100)
    fees_paid = random.choices(["Yes", "No"], weights=[0.78, 0.22])[0]
    fees_due = 0 if fees_paid == "Yes" else random.choice([5000, 7500, 10000, 12500])
    registered_for_drive = random.choices(["Yes", "No"], weights=[0.62, 0.38])[0]

    rows.append({
        "Student Name": name,
        "Roll No": roll,
        "Department": dept,
        "Semester": random.choice([3, 4, 5]),
        "Subject": subject,
        "Marks": marks,
        "Attendance %": attendance,
        "Fees Paid": fees_paid,
        "Fees Due (Rs)": fees_due,
        "Registered for Drive": registered_for_drive,
    })

# --- Intentional data-quality issues for the Clean & Anomalies demo ---

# 1. A duplicate row
rows.append(dict(rows[5]))

# 2. Missing attendance value
rows[10]["Attendance %"] = None

# 3. Missing fees status
rows[15]["Fees Paid"] = None

# 4. Attendance entered as >100% (bad data entry)
rows[20]["Attendance %"] = 114

# 5. A negative mark (impossible score)
rows[25]["Marks"] = -12

# 6. A blank student name
rows[30]["Student Name"] = None

# 7. Inconsistent casing / stray whitespace in a categorical field
rows[35]["Fees Paid"] = "yes "

random.shuffle(rows)

# --- Build workbook ---
wb = Workbook()
ws = wb.active
ws.title = "Academic Records"

headers = list(rows[0].keys())
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10.5)

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, row in enumerate(rows, start=2):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
        cell.font = BODY_FONT
        if header in ("Marks", "Attendance %", "Fees Due (Rs)", "Semester"):
            cell.alignment = Alignment(horizontal="center")

# Column widths
widths = {"A": 20, "B": 10, "C": 12, "D": 10, "E": 20, "F": 9, "G": 14, "H": 11,
          "I": 14, "J": 18}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

wb.save("demo_academic_data.xlsx")
print(f"Wrote demo_academic_data.xlsx with {len(rows)} rows")
